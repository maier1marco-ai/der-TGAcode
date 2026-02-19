# ==============================================================================
# der TGAcode – KI-gestützte Nachtragsprüfung (Streamlit)
# - Stammdaten/Gedächtnis pro Projekt (_projekt_stammdaten.txt)
# - Zwei-Agenten-Analyse (Analyst -> Fragen; Gutachter -> Prüfbericht)
# - Separater Schritt: strukturierte JSON-Zusammenfassung (Schema mit Fallback)
# - Excel-Deckblatt (.xlsx) befüllen via openpyxl (Upload + Repo-Fallback)
# - Robust gegen 429/Quota (Backoff) und 404/Not Found (Modellrotation)
# - Eco-Modus und Caching für Quota-Schonung
# - NEU: Korrekturen/Ergänzungen nach der Prüfung zur gezielten Überarbeitung
# ==============================================================================

import streamlit as st
import os
import time
import json
import hashlib
from io import BytesIO

# Bibliotheken prüfen und laden
try:
    from PyPDF2 import PdfReader
    import google.generativeai as genai
    import chromadb
    from sentence_transformers import SentenceTransformer
    import openpyxl
except ImportError as e:
    st.error(
        f"Eine benötigte Bibliothek fehlt: {e}. "
        "Bitte prüfe requirements.txt (streamlit, PyPDF2, google-generativeai, chromadb, sentence-transformers, openpyxl)."
    )
    st.stop()

# Seiten-Setup
st.set_page_config(page_title="der TGAcode", layout="wide")

# ==============================================================================
# Globale Konfiguration
# ==============================================================================
VAULT = "vault_tgacode"
try:
    os.makedirs(VAULT, exist_ok=True)
except Exception as e:
    st.error(f"Konnte das Verzeichnis '{VAULT}' nicht erstellen. Fehler: {e}")
    st.stop()

# Gemini API-Setup
try:
    api_key = st.secrets.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY nicht in Streamlit Secrets gefunden.")
    genai.configure(api_key=api_key)
except Exception as e:
    st.error(f"Fehler bei der Konfiguration des Gemini API Keys: {e}")
    st.stop()

# ==============================================================================
# Dynamische Modellauswahl & Backoff
# ==============================================================================

def discover_supported_models():
    """
    Fragt verfügbare Gemini-Modelle ab und filtert auf solche,
    die generateContent unterstützen. Liefert sortierte Liste.
    """
    supported = []
    try:
        for m in genai.list_models():
            name = getattr(m, "name", None)
            methods = getattr(m, "supported_generation_methods", []) or []
            if name and ("generateContent" in methods):
                supported.append(name)
    except Exception:
        # Fallback-Liste ohne potenziell eingeschränkte Modelle
        supported = [
            "gemini-2.0-flash",
            "gemini-1.5-flash",
            "gemini-1.5-pro",
            "gemini-1.0-pro",
        ]

    def sort_key(n: str):
        # Bevorzuge "flash", dann "pro", grob nach Major-Version absteigend
        tier = 0 if "flash" in n else (1 if "pro" in n else 2)
        import re
        m = re.search(r"gemini-(\d+)", n)
        major = int(m.group(1)) if m else 0
        return (tier, -major)

    return sorted(set(supported), key=sort_key)

@st.cache_resource
def get_models():
    """Instanziert nutzbare Modelle basierend auf dynamischer Discovery."""
    names = discover_supported_models()
    instances = []
    for name in names:
        try:
            instances.append(genai.GenerativeModel(name))
        except Exception:
            continue
    return instances, names

def generate_with_backoff(prompt, max_output_tokens=1536, temperature=0.3, attempts_per_model=2):
    """
    generateContent mit Rotation bei 404/Not Found und Backoff bei 429/Quota.
    Gibt resp.text zurück.
    """
    models, names = get_models()
    if not models:
        raise RuntimeError("Keine nutzbaren Gemini-Modelle gefunden. Prüfe API-Zugang/Billing.")

    last_err = None
    for idx, model in enumerate(models):
        model_name = names[idx]
        for i in range(attempts_per_model):
            try:
                resp = model.generate_content(
                    prompt,
                    generation_config={
                        "max_output_tokens": max_output_tokens,
                        "temperature": temperature,
                    },
                )
                st.sidebar.caption(f"KI-Modell verwendet: {model_name}")
                return resp.text
            except Exception as e:
                msg = str(e).lower()
                if "404" in msg or "not found" in msg:
                    last_err = e
                    break  # direkt nächstes Modell
                elif "429" in msg or "quota" in msg:
                    delay = min(2 ** i, 8)
                    time.sleep(delay)
                    last_err = e
                    continue
                else:
                    last_err = e
                    break
    raise last_err if last_err else RuntimeError("KI-Generierung fehlgeschlagen.")

def summary_json_schema():
    # Schlankes, kompatibles Schema – ohne additionalProperties
    return {
        "type": "object",
        "properties": {
            "vob_check": {"type": "string"},
            "technische_pruefung": {"type": "string"},
            "preis_check": {"type": "string"},
            "gesamtsumme_korrigiert": {"type": "string"},
            "empfehlung": {"type": "string"},
            "naechste_schritte": {"type": "string"},
        },
        "required": [
            "vob_check",
            "technische_pruefung",
            "preis_check",
            "gesamtsumme_korrigiert",
            "empfehlung",
            "naechste_schritte",
        ]
    }

def generate_json_with_backoff(prompt, json_schema=None, attempts_per_model=2):
    """
    Erzeugt JSON via response_mime_type=application/json.
    1) Versucht Schema-Modus (wenn json_schema übergeben wird).
    2) Bei Schema-Fehlern oder Nichtunterstützung: Fallback ohne Schema
       + strikte Prompt-Vorgabe "Nur ein JSON-Objekt zurückgeben".
    Gibt ein Python-Dict zurück.
    """
    models, names = get_models()
    if not models:
        raise RuntimeError("Keine nutzbaren Gemini-Modelle gefunden. Prüfe API-Zugang/Billing.")

    def try_call(model, use_schema=True):
        if use_schema and json_schema:
            return model.generate_content(
                prompt,
                generation_config={
                    "response_mime_type": "application/json",
                    "response_schema": json_schema,
                    "temperature": 0.2,
                    "max_output_tokens": 512,
                },
            )
        else:
            strict_prompt = (
                prompt
                + "\n\nGib ausschließlich ein einzelnes, valides JSON-Objekt zurück – ohne Erklärtext, keine Code-Fences."
            )
            return model.generate_content(
                strict_prompt,
                generation_config={
                    "response_mime_type": "application/json",
                    "temperature": 0.2,
                    "max_output_tokens": 512,
                },
            )

    last_err = None
    for idx, model in enumerate(models):
        model_name = names[idx]
        # Zweiphasig: (1) mit Schema, (2) ohne Schema
        for phase in (True, False):
            for i in range(attempts_per_model):
                try:
                    resp = try_call(model, use_schema=phase)
                    st.sidebar.caption(f"KI (JSON): {model_name} | Schema={'on' if phase else 'off'}")
                    return json.loads(resp.text)
                except Exception as e:
                    msg = str(e).lower()
                    # 404 → direkt nächstes Modell
                    if "404" in msg or "not found" in msg:
                        last_err = e
                        break
                    # Schema nicht unterstützt → sofort in nächste Phase (ohne Schema)
                    if phase and ("schema" in msg or "unknown field" in msg or "unsupported" in msg):
                        last_err = e
                        break
                    # 429 / Quota → Backoff
                    if "429" in msg or "quota" in msg:
                        delay = min(2 ** i, 8)
                        time.sleep(delay)
                        last_err = e
                        continue
                    # Anderes Problem → nächster Versuch mit nächstem Modell/Phase
                    last_err = e
                    break
            # Falls 404 oder Schema-Problem: Phase/Modell wechseln
            if last_err and ("404" in str(last_err) or "schema" in str(last_err).lower() or "unknown field" in str(last_err).lower()):
                continue
    raise last_err if last_err else RuntimeError("Strukturierte JSON-Generierung fehlgeschlagen.")

# ==============================================================================
# Helferfunktionen & Vektorindex
# ==============================================================================

@st.cache_resource
def get_embedder():
    try:
        return SentenceTransformer("all-MiniLM-L6-v2")
    except Exception as e:
        st.error(f"Fehler beim Laden des Embedding-Modells: {e}")
        st.stop()

def read_pdf(file):
    """Extrahiert Text aus PDF-Datei."""
    text = ""
    try:
        reader = PdfReader(file)
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    except Exception:
        pass
    return text

def index_project(path, p_id, embedder, chroma_client):
    """Zerlegt PDFs in Chunks, berechnet Embeddings und legt sie in ChromaDB ab."""
    col = chroma_client.get_or_create_collection(p_id)
    ids = col.get().get("ids", [])
    if ids:
        col.delete(ids=ids)
    for f in os.listdir(path):
        if f.lower().endswith(".pdf"):
            text = read_pdf(os.path.join(path, f))
            words = text.split()
            chunks = [" ".join(words[i:i+400]) for i in range(0, len(words), 400)]
            if chunks:
                embeddings = [embedder.encode(c).tolist() for c in chunks]
                col.add(
                    ids=[f"{f}_{i}" for i in range(len(chunks))],
                    documents=chunks,
                    embeddings=embeddings,
                )

# Modelle/Clients laden
embedder = get_embedder()
chroma_client = chromadb.Client()

# ==============================================================================
# UI-Design (CSS)
# ==============================================================================
UI_CSS = """
<style>
    body { color: #fafafa; background-color: #0d1117; }
    .stApp { background-color: #0d1117; }
    h1, h2, h3 { color: #c9d1d9; }
    .top-nav {
        background-color: #161b22;
        padding: 1rem 2rem;
        border-bottom: 2px solid #00f2fe;
        margin-bottom: 2rem;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .logo { font-size: 26px; font-weight: 800; color: #f0f6fc; }
    .accent { color: #00f2fe; }
    .stButton>button {
        background: linear-gradient(45deg, #00f2fe, #2c7fff);
        color: white; border: none; width: 100%;
        font-weight: bold; padding: 10px 0; border-radius: 8px;
        transition: transform 0.1s ease-in-out;
    }
    .stButton>button:hover {
        transform: scale(1.02);
        box-shadow: 0 0 15px #00f2fe;
    }
    .report-box {
        background-color: #161b22;
        padding: 25px; border-radius: 10px;
        border: 1px solid #30363d; line-height: 1.6;
    }
    .report-box h1, .report-box h3 {
        border-bottom: 1px solid #30363d; padding-bottom: 8px;
    }
</style>
<div class="top-nav">
    <div class="logo">der <span class="accent">TGAcode</span></div>
    <div style="color: #8b949e;">AI-Powered Project Analysis</div>
</div>
"""

# ==============================================================================
# Haupt-App
# ==============================================================================
def main():
    st.markdown(UI_CSS, unsafe_allow_html=True)

    # Eco-Modus: weniger KI-Aufrufe (überspringt Fragen-Agent)
    eco_mode = st.sidebar.toggle(
        "Eco-Modus (Quota-schonend)", value=False,
        help="Reduziert KI-Aufrufe: Fragen-Phase wird übersprungen, Kontext via Nachtragstext."
    )

    st.header("Projektauswahl")
    c1, c2 = st.columns([1, 2])

    with c1:
        firmen = [f for f in os.listdir(VAULT) if os.path.isdir(os.path.join(VAULT, f))]
        sel_f = st.selectbox("Firma auswählen", ["--"] + firmen, label_visibility="collapsed")

        projekte = []
        if sel_f != "--":
            projekte = [p for p in os.listdir(os.path.join(VAULT, sel_f))]
        sel_p = st.selectbox("Projekt auswählen", ["--"] + projekte, label_visibility="collapsed")

    with c2:
        with st.expander("➕ Neues Projekt oder Firma anlegen"):
            nc1, nc2 = st.columns(2)
            with nc1:
                nf = st.text_input("Neue Firma")
                if st.button("Firma anlegen"):
                    if nf.strip():
                        os.makedirs(os.path.join(VAULT, nf.strip()), exist_ok=True)
                        st.rerun()
            with nc2:
                np_firma = st.selectbox("Für Firma", ["--"] + firmen)
                np = st.text_input("Neues Projekt")
                if st.button("Projekt anlegen") and np_firma != "--":
                    if np.strip():
                        os.makedirs(os.path.join(VAULT, np_firma, np.strip()), exist_ok=True)
                        st.rerun()

    st.markdown("---")

    if sel_f != "--" and sel_p != "--":
        p_path = os.path.join(VAULT, sel_f, sel_p)
        p_id = f"{sel_f}_{sel_p}".replace(" ", "_")

        st.header(f"Projekt-Dashboard: {sel_p}")
        t1, t2 = st.tabs(["📁 Projekt-Akte", "🚀 Nachtrags-Prüfung"])

        # Tab 1 – Projekt-Akte
        with t1:
            st.subheader("Stammdaten & Projekt-Regeln (Gedächtnis)")
            stammdaten_path = os.path.join(p_path, "_projekt_stammdaten.txt")
            current_stammdaten = ""
            if os.path.exists(stammdaten_path):
                with open(stammdaten_path, "r", encoding="utf-8") as f:
                    current_stammdaten = f.read()

            stammdaten_input = st.text_area(
                "Permanente Regeln/Absprachen (z. B. 'Stundensatz Fa. Reiter ist 48 €').",
                value=current_stammdaten, height=150
            )
            if st.button("Stammdaten speichern"):
                os.makedirs(p_path, exist_ok=True)
                with open(stammdaten_path, "w", encoding="utf-8") as f:
                    f.write(stammdaten_input)
                st.success("Stammdaten wurden gespeichert!")

            st.markdown("---")
            st.subheader("Dokumente verwalten")
            up = st.file_uploader("Neue Dokumente hochladen", accept_multiple_files=True, type="pdf")
            if up:
                if st.button("In Akte speichern"):
                    os.makedirs(p_path, exist_ok=True)
                    for f in up:
                        with open(os.path.join(p_path, f.name), "wb") as o:
                            o.write(f.getbuffer())
                    st.rerun()

            st.markdown("---")
            col_a, col_b = st.columns([2, 1])
            with col_a:
                st.subheader("Bestehende Dokumente")
                docs = [d for d in os.listdir(p_path) if not d.startswith("_")] if os.path.isdir(p_path) else []
                if not docs:
                    st.info("Noch keine Dokumente in dieser Akte.")
                else:
                    for d in docs:
                        st.code(d)
            with col_b:
                st.subheader("Projekt-Wissen")
                if st.button("📚 Wissen neu indexieren"):
                    with st.spinner("Projektwissen wird analysiert und indexiert..."):
                        index_project(p_path, p_id, embedder, chroma_client)
                    st.success("Projektwissen ist auf dem neuesten Stand!")

        # Tab 2 – Nachtrags-Prüfung
        with t2:
            st.subheader("Nachtrag zur Prüfung hochladen")
            nt = st.file_uploader("Nachtrag PDF", accept_multiple_files=True, type="pdf", label_visibility="collapsed")

            if st.button("🔥 KI-Prüfung starten", type="primary"):
                if not nt:
                    st.warning("Bitte zuerst einen Nachtrag hochladen.")
                else:
                    with st.status("Starte KI-Analyse...", expanded=True) as status:
                        status.write("Vorbereitung…")
                        nt_text = "".join([read_pdf(f) for f in nt])
                        nt_hash = hashlib.sha256(nt_text.encode("utf-8")).hexdigest()

                        # Session-Caches zur Quota-Reduzierung
                        if "cache_questions" not in st.session_state:
                            st.session_state.cache_questions = {}
                        if "cache_report" not in st.session_state:
                            st.session_state.cache_report = {}

                        # Agent 1: Analyst (optional, via Eco-Modus)
                        questions = []
                        if not eco_mode:
                            status.write("Agent 1 (Analyst): Untersucht den Nachtrag…")
                            question_prompt = (
                                "Du bist ein Analyst für TGA-Bauprojekte. Lies den Nachtrag, "
                                "identifiziere die 3-5 Kernforderungen und formuliere für jede eine präzise Frage, "
                                "um relevante Infos in den Projektunterlagen zu finden. Gib NUR die Liste der Fragen aus.\n\n"
                                f"NACHTRAG:\n{nt_text[:4000]}"
                            )
                            try:
                                if nt_hash in st.session_state.cache_questions:
                                    questions = st.session_state.cache_questions[nt_hash]
                                else:
                                    q_text = generate_with_backoff(question_prompt, max_output_tokens=512, temperature=0.2)
                                    questions = [q.strip() for q in q_text.strip().split("\n") if q.strip()]
                                    st.session_state.cache_questions[nt_hash] = questions
                                status.update(label="Agent 1 (Analyst): Rechercheplan erstellt! ✅")
                            except Exception:
                                status.update(label="Agent 1: Fragengenerierung fehlgeschlagen – Eco-Fallback aktiv", state="error")
                                questions = []

                        # Agent 2: Kontextbeschaffung
                        status.write("Agent 2 (Gutachter): Sucht relevante Projektdaten…")
                        final_ctx = ""
                        try:
                            collection = chroma_client.get_or_create_collection(p_id)
                            if questions:
                                for q in questions:
                                    q_vec = embedder.encode(q).tolist()
                                    res = collection.query(query_embeddings=[q_vec], n_results=3)
                                    docs_block = "\n".join(res.get("documents", [[]])[0]) if res.get("documents") else ""
                                    final_ctx += f"Recherche-Ergebnis für Frage '{q}':\n{docs_block}\n\n---\n\n"
                            else:
                                # Eco-/Fallback: nutze Nachtragstext als Query
                                q_vec = embedder.encode(nt_text[:1000]).tolist()
                                res = collection.query(query_embeddings=[q_vec], n_results=5)
                                docs_block = "\n".join(res.get("documents", [[]])[0]) if res.get("documents") else ""
                                final_ctx += f"Kontext (Eco/Fallback):\n{docs_block}\n\n---\n\n"
                            status.update(label="Agent 2 (Gutachter): Daten aus Projekt-Akte geladen! ✅")
                        except Exception as e:
                            final_ctx = f"Fehler bei der Datenbeschaffung: {e}"
                            status.update(label="Agent 2: Kontextbeschaffung fehlgeschlagen", state="error")

                        # Für spätere Überarbeitungen merken:
                        st.session_state.current_nt_text = nt_text
                        st.session_state.current_final_ctx = final_ctx
                        st.session_state.current_stammdaten_text = ""
                        if os.path.exists(os.path.join(p_path, "_projekt_stammdaten.txt")):
                            with open(os.path.join(p_path, "_projekt_stammdaten.txt"), "r", encoding="utf-8") as f:
                                st.session_state.current_stammdaten_text = f.read()

                        # Agent 2: Finaler Bericht (Markdown)
                        status.write("Agent 2 (Gutachter): Erstellt den finalen Bericht…")
                        report_prompt = f"""
                        SYSTEM: Du bist 'der TGAcode', ein KI-Gutachter für TGA-Bauprojekte (VOB).
                        DEINE AUFGABE: Erstelle einen finalen Prüfbericht im Markdown-Format mit:
                        - Zusammenfassung (3-4 Stichpunkte)
                        - VOB/B-Konformitäts-Check
                        - Technische Prüfung & Preis-Check
                        - Empfehlung & Nächste Schritte

                        PROJEKT-STAMMDATEN (höchste Priorität):
                        ---
                        {st.session_state.current_stammdaten_text}
                        ---

                        DER ZU PRÜFENDE NACHTRAG:
                        ---
                        {st.session_state.current_nt_text}
                        ---

                        RECHERCHE-ERGEBNISSE AUS DER PROJEKT-AKTE:
                        ---
                        {st.session_state.current_final_ctx}
                        ---
                        """
                        try:
                            if nt_hash in st.session_state.cache_report:
                                st.session_state.report = st.session_state.cache_report[nt_hash]
                            else:
                                st.session_state.report = generate_with_backoff(
                                    report_prompt, max_output_tokens=1400, temperature=0.2
                                )
                                st.session_state.cache_report[nt_hash] = st.session_state.report
                        except Exception as e:
                            status.update(label=f"Berichtserstellung fehlgeschlagen: {e}", state="error")
                            st.session_state.report = ""

                        # Separater Schritt: strukturierte JSON-Zusammenfassung
                        status.write("Agent 2 (Gutachter): Erstellt die strukturierte Zusammenfassung (JSON)…")
                        json_prompt = f"""
                        Erzeuge eine komprimierte, sachliche JSON-Zusammenfassung der Prüfung
                        mit den Feldern: vob_check, technische_pruefung, preis_check,
                        gesamtsumme_korrigiert, empfehlung, naechste_schritte.

                        Nutze ausschließlich diese Quellen:
                        1) Projekt-Stammdaten:
                        {st.session_state.current_stammdaten_text}

                        2) Nachtrag (Volltext; ggf. gekürzt):
                        {st.session_state.current_nt_text[:10000]}

                        3) Recherchierte Projekt-Kontexte (gekürzt):
                        {st.session_state.current_final_ctx[:10000]}

                        4) Eigener Bericht (Auszug):
                        {st.session_state.report[:4000]}

                        Formuliere kurze, klare Werte. Keine Erläuterung, nur die reinen Feldwerte.
                        """
                        st.session_state.json_prompt = json_prompt  # für spätere Regeneration/Korrekturen speichern
                        try:
                            st.session_state.summary = generate_json_with_backoff(
                                json_prompt, summary_json_schema()
                            )
                            status.update(label="Analyse abgeschlossen!", state="complete", expanded=False)
                        except Exception as e:
                            status.update(label=f"JSON-Erstellung fehlgeschlagen: {e}", state="error")
                            st.session_state.summary = None

            # Berichtanzeige + JSON-Status
            if "report" in st.session_state:
                st.markdown("---")
                st.subheader("Ergebnis der KI-Prüfung")
                st.markdown(f"<div class='report-box'>{st.session_state.report}</div>", unsafe_allow_html=True)

                st.markdown("#### Strukturierte Zusammenfassung (JSON)")
                if st.session_state.get("summary") is None:
                    st.warning("Noch keine JSON-Zusammenfassung vorhanden.")
                    if st.session_state.get("json_prompt"):
                        if st.button("JSON-Zusammenfassung jetzt erzeugen (erneut)"):
                            try:
                                st.session_state.summary = generate_json_with_backoff(
                                    st.session_state.json_prompt,
                                    summary_json_schema()
                                )
                                st.success("JSON-Zusammenfassung erzeugt.")
                            except Exception as e:
                                st.error(f"JSON-Erstellung fehlgeschlagen: {e}")
                else:
                    st.code(json.dumps(st.session_state.summary, ensure_ascii=False, indent=2), language="json")

                # NEU: Korrekturen/Ergänzungen nach der Prüfung
                st.markdown("---")
                st.subheader("Korrekturen/Ergänzungen an der Prüfung")
                corrections = st.text_area(
                    "Teile der KI konkrete Änderungen mit (z. B. 'Stundenlohn Fa. Reiter ist 48 €, Position 3.1 Menge 12 statt 10').",
                    placeholder="Kurze, präzise Hinweise eintragen..."
                )
                if st.button("Bericht und Zusammenfassung mit Korrekturen überarbeiten"):
                    if corrections.strip():
                        with st.spinner("Überarbeitung läuft…"):
                            # Bericht verfeinern
                            refine_report_prompt = f"""
                            SYSTEM: Du bist 'der TGAcode', KI-Gutachter.
                            Überarbeite den bestehenden Bericht sachlich und präzise anhand der Korrekturen des Nutzers.
                            Behalte die gleiche Gliederung (Zusammenfassung, VOB-Check, Technik/Preis-Check, Empfehlung).

                            Bestehender Bericht:
                            ---
                            {st.session_state.report[:6000]}
                            ---

                            Korrekturen des Nutzers:
                            ---
                            {corrections}
                            ---

                            Zusätzlicher Kontext (falls nötig):
                            - Stammdaten: {st.session_state.get('current_stammdaten_text', '')[:2000]}
                            - Nachtrag (Kurzfassung): {st.session_state.get('current_nt_text', '')[:3000]}
                            - Recherche-Kontext: {st.session_state.get('current_final_ctx', '')[:3000]}
                            """
                            try:
                                st.session_state.report = generate_with_backoff(
                                    refine_report_prompt, max_output_tokens=1400, temperature=0.2
                                )
                                # JSON anhand des neuen Berichts und Korrekturen neu erzeugen
                                refined_json_prompt = f"""
                                Erzeuge eine komprimierte, sachliche JSON-Zusammenfassung der Prüfung
                                mit den Feldern: vob_check, technische_pruefung, preis_check,
                                gesamtsumme_korrigiert, empfehlung, naechste_schritte.

                                Quellen:
                                1) Projekt-Stammdaten:
                                {st.session_state.get('current_stammdaten_text', '')}

                                2) Nachtrag (Volltext; ggf. gekürzt):
                                {st.session_state.get('current_nt_text', '')[:10000]}

                                3) Recherchierte Projekt-Kontexte (gekürzt):
                                {st.session_state.get('current_final_ctx', '')[:10000]}

                                4) Überarbeiteter Bericht (Auszug):
                                {st.session_state.report[:4000]}

                                5) Korrekturen des Nutzers:
                                {corrections}

                                Keine Erläuterung, nur reine Feldwerte im JSON-Objekt.
                                """
                                st.session_state.summary = generate_json_with_backoff(
                                    refined_json_prompt, summary_json_schema()
                                )
                                st.success("Bericht und JSON-Zusammenfassung wurden überarbeitet.")
                            except Exception as e:
                                st.error(f"Überarbeitung fehlgeschlagen: {e}")
                    else:
                        st.warning("Bitte konkrete Korrekturen/Ergänzungen eintragen.")

                # Deckblatt aus Excel-Vorlage – Upload + Repo-Fallback
                st.markdown("---")
                st.subheader("Deckblatt aus Excel-Vorlage erstellen")

                template_file = st.file_uploader(
                    "Excel-Deckblatt hochladen (.xlsx bevorzugt)",
                    type=None,  # akzeptiert alles; wir prüfen die Endung selbst
                    accept_multiple_files=False,
                    help="Falls Upload blockiert ist, nutze die Vorlagen-Auswahl aus dem Repository unten."
                )

                repo_templates_dir = "templates"
                available_repo_templates = []
                if os.path.isdir(repo_templates_dir):
                    available_repo_templates = [
                        f for f in os.listdir(repo_templates_dir)
                        if f.lower().endswith(".xlsx")
                    ]

                use_repo_template = False
                selected_repo_template = None
                if available_repo_templates:
                    st.info("Alternativ Vorlage direkt aus dem Repository wählen.")
                    selected_repo_template = st.selectbox(
                        "Vorlage aus Repository wählen", ["--"] + available_repo_templates
                    )
                    use_repo_template = (selected_repo_template and selected_repo_template != "--")

                report_data = st.session_state.get("summary")

                if report_data:
                    workbook = None

                    # A) Upload bevorzugt – akzeptiere nur .xlsx nach Endung
                    if template_file is not None:
                        st.caption(
                            f"Upload erkannt: name={template_file.name}, mime={getattr(template_file, 'type', 'unbekannt')}"
                        )
                        if template_file.name.lower().endswith(".xlsx"):
                            try:
                                workbook = openpyxl.load_workbook(template_file)
                            except Exception as e:
                                st.error(f"Excel konnte nicht geladen werden: {e}")
                        else:
                            st.warning("Bitte eine .xlsx-Datei hochladen (Excel-OpenXML-Format).")

                    # B) Fallback: Vorlage aus Repository laden
                    if workbook is None and use_repo_template:
                        try:
                            wb_path = os.path.join(repo_templates_dir, selected_repo_template)
                            workbook = openpyxl.load_workbook(wb_path)
                            st.caption(f"Vorlage aus Repository geladen: {selected_repo_template}")
                        except Exception as e:
                            st.error(f"Vorlage aus Repository konnte nicht geladen werden: {e}")

                    # C) Befüllen & Download
                    if workbook is None:
                        st.info("Keine Excel-Vorlage verfügbar. Bitte .xlsx hochladen oder Vorlage aus Repository wählen.")
                    else:
                        sheet = workbook.active
                        placeholders = {
                            "[VOB_CHECK]": str(report_data.get("vob_check", "")),
                            "[TECHNISCHE_PRUEFUNG]": str(report_data.get("technische_pruefung", "")),
                            "[PREIS_CHECK]": str(report_data.get("preis_check", "")),
                            "[GESAMTSUMME_KORRIGIERT]": str(report_data.get("gesamtsumme_korrigiert", "")),
                            "[EMPFEHLUNG]": str(report_data.get("empfehlung", "")),
                            "[NAECHSTE_SCHRITTE]": str(report_data.get("naechste_schritte", "")),
                        }

                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    if cell.value in placeholders:
                                        cell.value = placeholders[cell.value]

                        output_stream = BytesIO()
                        try:
                            workbook.save(output_stream)
                            output_stream.seek(0)
                            st.success("Excel-Vorlage erfolgreich befüllt!")
                            out_name = template_file.name if template_file else selected_repo_template
                            st.download_button(
                                label="✅ Fertiges Deckblatt herunterladen",
                                data=output_stream,
                                file_name=f"Deckblatt_{sel_p}_{out_name}",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        except Exception as e:
                            st.error(f"Speichern der Excel-Ausgabe ist fehlgeschlagen: {e}")

# Start
if __name__ == "__main__":
    main()
